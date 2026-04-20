import csv
from decimal import Decimal
from io import BytesIO
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.dateparse import parse_date

from .forms import FormularioSolicitudTiquete, FormularioRegistroPago, FormularioRegistroPagoAdmin, FormularioInventario, FormularioAumentarInventario, FormularioEditarConsumo
from .models import Consumo, ConsumoLog, SolicitudTiquete, InventarioTiquetes, RegistroPago
from users.models import Empleado, Administrador

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover - fallback for environments without openpyxl
    Workbook = None
    Alignment = Font = PatternFill = get_column_letter = None

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
except ImportError:  # pragma: no cover - fallback for environments without reportlab
    colors = letter = getSampleStyleSheet = None
    Paragraph = SimpleDocTemplate = Spacer = Table = TableStyle = None

# ==========================================
# VISTAS DE EMPLEADO
# ==========================================

@login_required
def solicitar_tiquete(request):
    from django.db import transaction
    try:
        empleado = request.user.empleado_perfil
    except Empleado.DoesNotExist:
        messages.error(request, "Solo los empleados pueden solicitar tiquetes.")
        return redirect("dashboard_empleado")

    # Obtener información básica para mostrar en el formulario (GET o POST fallido)
    hoy = timezone.localdate()
    inventario_info = InventarioTiquetes.objects.filter(mes__month=hoy.month, mes__year=hoy.year).first()
    max_permitido = inventario_info.max_tiquetes_por_empleado if inventario_info else 20
    stock_disponible = inventario_info.cantidad_disponible if inventario_info else 0

    if request.method == "POST":
        formulario = FormularioSolicitudTiquete(request.POST, empleado=empleado)
        if formulario.is_valid():
            cantidad = formulario.cleaned_data["cantidad"]
            solicitud = formulario.save(commit=False)
            solicitud.empleado = empleado

            try:
                with transaction.atomic():
                    # Obtener inventario actual con bloqueo para actualización (para la lógica de aprobación)
                    inventario = InventarioTiquetes.objects.filter(
                        mes__month=hoy.month, mes__year=hoy.year
                    ).select_for_update().first()
                    
                    if not inventario:
                        messages.error(request, "No hay inventario configurado para este periodo.")
                        return render(request, "schedule/solicitar_tiquete.html", {
                            "formulario": formulario,
                            "max_permitido": max_permitido,
                            "stock_disponible": stock_disponible
                        })
                    
                    # Usamos los datos actualizados del bloqueo para la lógica
                    stock_real = inventario.cantidad_disponible
                    limite_real = inventario.max_tiquetes_por_empleado

                    # 1. Validar Inventario Global
                    if cantidad > stock_real:
                        solicitud.estado = "rechazado"
                        solicitud.save()
                        messages.error(request, f"Solicitud rechazada por falta de stock global. (Disponible: {stock_real})")
                        return redirect("dashboard_empleado")
                    
                    # 2. Validar Límite por Empleado
                    tiquetes_ya_aprobados = SolicitudTiquete.objects.filter(
                        empleado=empleado, 
                        estado="aprobado",
                        fecha_solicitud__month=hoy.month,
                        fecha_solicitud__year=hoy.year
                    ).aggregate(total=Sum("cantidad"))["total"] or 0
                    
                    if (tiquetes_ya_aprobados + cantidad) > limite_real:
                        solicitud.estado = "rechazado"
                        solicitud.save()
                        messages.error(request, f"Solicitud rechazada. Has excedido tu límite mensual de {limite_real} tiquetes.")
                        return redirect("dashboard_empleado")

                    # Si todo está bien, aprobamos automáticamente
                    inventario.cantidad_disponible -= cantidad
                    inventario.save()
                    
                    solicitud.estado = "aprobado"
                    solicitud.save()
                    
                    messages.success(request, f"¡Éxito! Tu tiquete ha sido otorgado automáticamente.")
                    return redirect("dashboard_empleado")
            except Exception as e:
                messages.error(request, f"Error al procesar la solicitud: {str(e)}")
                return render(request, "schedule/solicitar_tiquete.html", {
                    "formulario": formulario,
                    "max_permitido": max_permitido,
                    "stock_disponible": stock_disponible
                })
    else:
        formulario = FormularioSolicitudTiquete(empleado=empleado)

    return render(request, "schedule/solicitar_tiquete.html", {
        "formulario": formulario,
        "max_permitido": max_permitido,
        "stock_disponible": stock_disponible
    })


@login_required
def registrar_pago(request):
    try:
        empleado = request.user.empleado_perfil
    except Empleado.DoesNotExist:
        return redirect("dashboard_empleado")

    if request.method == "POST":
        formulario = FormularioRegistroPago(request.POST)
        if formulario.is_valid():
            pago = formulario.save(commit=False)
            pago.empleado = empleado
            pago.save()
            messages.success(request, "Pago registrado. Pendiente de validación por Gestión Humana.")
            return redirect("dashboard_empleado")
    else:
        formulario = FormularioRegistroPago()

    return render(request, "schedule/registrar_pago.html", {"formulario": formulario})


@login_required
def consultar_estado_cuenta(request):
    try:
        empleado = request.user.empleado_perfil
    except Empleado.DoesNotExist:
        return redirect("dashboard_empleado")

    # Filtrado por mes en el historial de compras
    hoy = timezone.localdate()
    try:
        mes_actual = int(request.GET.get('mes', hoy.month))
        anio_actual = int(request.GET.get('anio', hoy.year))
    except ValueError:
        mes_actual = hoy.month
        anio_actual = hoy.year
        
    # Validar rangos de fecha
    if mes_actual < 1 or mes_actual > 12:
        mes_actual = hoy.month
        anio_actual = hoy.year

    # Cálculos para la UI de navegación de meses
    next_month = mes_actual + 1 if mes_actual < 12 else 1
    next_year = anio_actual if mes_actual < 12 else anio_actual + 1
    prev_month = mes_actual - 1 if mes_actual > 1 else 12
    prev_year = anio_actual if mes_actual > 1 else anio_actual - 1
    
    nombres_meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    mes_nombre = nombres_meses[mes_actual - 1]

    # Historial de compras (tiquetes aprobados) del mes especificado
    compras = SolicitudTiquete.objects.filter(
        empleado=empleado, 
        estado="aprobado",
        fecha_solicitud__year=anio_actual,
        fecha_solicitud__month=mes_actual
    ).order_by("-fecha_solicitud")
    
    # Obtener precio actual del inventario (solo para propósitos informativos generales)
    inventario = InventarioTiquetes.objects.order_by("-mes").first()
    precio_tiquete = inventario.precio_tiquete if inventario else Decimal("10000.00")

    # Calcular valor detallado usando el precio histórico grabado (precio_unitario)
    for compra in compras:
        compra.valor_total = compra.cantidad * getattr(compra, 'precio_unitario', Decimal('10000.00'))
        
    # Historial de pagos validados
    pagos = RegistroPago.objects.filter(empleado=empleado, validado_por_gh=True).order_by("-fecha_pago")
    
    tiquetes_aprobados = compras.aggregate(total=Sum("cantidad"))["total"] or 0
    tiquetes_consumidos = Consumo.objects.filter(
        empleado=empleado, 
        fecha_consumo__month=mes_actual,
        fecha_consumo__year=anio_actual
    ).count()
    tiquetes_disponibles = max(0, tiquetes_aprobados - tiquetes_consumidos)
    
    total_pagos_validados = pagos.aggregate(total=Sum("valor_pagado"))["total"] or Decimal("0.00")
    
    from django.db.models import F
    deuda_total = SolicitudTiquete.objects.filter(empleado=empleado, estado="aprobado").annotate(
        costo_total=F('cantidad') * F('precio_unitario')
    ).aggregate(total=Sum('costo_total'))["total"] or Decimal("0.00")
    
    saldo_pendiente = deuda_total - total_pagos_validados

    return render(
        request,
        "schedule/consultar_estado_cuenta.html",
        {
            "empleado": empleado,
            "compras": compras,
            "pagos": pagos,
            "saldo_pendiente": saldo_pendiente,
            "tiquetes_comprados": tiquetes_aprobados,
            "tiquetes_disponibles": tiquetes_disponibles,
            "precio_tiquete": precio_tiquete,
            "ultima_actualizacion": timezone.now(),
            "mes_nombre": mes_nombre,
            "anio_actual": anio_actual,
            "prev_month": prev_month,
            "prev_year": prev_year,
            "next_month": next_month,
            "next_year": next_year,
        },
    )

@login_required
def confirmar_pago_empleado(request, pago_id):
    try:
        empleado = request.user.empleado_perfil
    except Empleado.DoesNotExist:
        return redirect("dashboard_empleado")

    pago = get_object_or_404(RegistroPago, id=pago_id, empleado=empleado)
    pago.confirmado_por_empleado = True
    pago.save()
    messages.success(request, f"Pago de ${pago.valor_pagado:.0f} confirmado exitosamente.")
    return redirect("consultar_estado_cuenta")


# ==========================================
# VISTAS DE ADMINISTRADOR
# ==========================================

@login_required
def gestionar_solicitud(request, solicitud_id, accion):
    if request.user.role != 'administrador' and not request.user.is_superuser:
        return redirect("dashboard_empleado")

    solicitud = get_object_or_404(SolicitudTiquete, id=solicitud_id)
    if accion == "aprobar":
        inventario = InventarioTiquetes.objects.order_by("-mes").first()
        if inventario and inventario.cantidad_disponible >= solicitud.cantidad:
            inventario.cantidad_disponible -= solicitud.cantidad
            inventario.save()
            solicitud.estado = "aprobado"
            messages.success(request, f"Solicitud de {solicitud.empleado} aprobada e inventario actualizado.")
        else:
            messages.error(request, "No hay suficiente inventario para aprobar esta solicitud.")
            return redirect("dashboard_admin")
    elif accion == "rechazar":
        solicitud.estado = "rechazado"
        messages.info(request, f"Solicitud de {solicitud.empleado} rechazada.")
    
    solicitud.save()
    return redirect("dashboard_admin")


@login_required
def registrar_pago_efectivo(request):
    if request.user.role != 'administrador' and not request.user.is_superuser:
        return redirect("dashboard_empleado")

    if request.method == "POST":
        formulario = FormularioRegistroPagoAdmin(request.POST)
        if formulario.is_valid():
            pago = formulario.save(commit=False)
            pago.validado_por_gh = True
            pago.confirmado_por_empleado = False
            pago.comprobante = "Pago en efectivo (Caja RRHH)"
            pago.save()
            messages.success(request, f"Pago de ${pago.valor_pagado:.0f} registrado para {pago.empleado.user.get_full_name() or pago.empleado.user.username}.")
            return redirect("dashboard_admin")
    else:
        formulario = FormularioRegistroPagoAdmin()

    return render(request, "schedule/registrar_pago_efectivo.html", {"formulario": formulario})


@login_required
def gestionar_inventario(request):
    if request.user.role != 'administrador' and not request.user.is_superuser:
        return redirect("dashboard_empleado")

    hoy = timezone.now()
    # Buscar si ya existe un inventario para este mes y año
    inventario = InventarioTiquetes.objects.filter(
        mes__month=hoy.month, 
        mes__year=hoy.year
    ).first()
    
    # Si no hay uno para este mes, buscar el último de meses anteriores para precargar datos
    if not inventario:
        ultimo_global = InventarioTiquetes.objects.order_by("-mes").first()
        if ultimo_global:
            from datetime import date
            instancia_inicial = InventarioTiquetes(
                mes=date(hoy.year, hoy.month, 1),
                cantidad_inicial=ultimo_global.cantidad_inicial,
                max_tiquetes_por_empleado=ultimo_global.max_tiquetes_por_empleado,
                precio_tiquete=ultimo_global.precio_tiquete
            )
        else:
            instancia_inicial = None
    else:
        instancia_inicial = inventario

    ya_existe = inventario is not None

    if request.method == "POST":
        formulario = FormularioInventario(request.POST, instance=instancia_inicial, ya_existe=ya_existe)
        if formulario.is_valid():
            inv = formulario.save(commit=False)
            
            # Si es una creación nueva para el mes
            if not ya_existe:
                inv.cantidad_disponible = inv.cantidad_inicial
                inv.save()
                messages.success(request, f"Inventario para {inv.mes.strftime('%B %Y')} configurado con éxito.")
            else:
                inv.save()
                messages.success(request, "Límites y precios actualizados correctamente.")
            
            return redirect("dashboard_admin")
    else:
        # Si es nuevo mes, pasamos los valores del último pero sin ser la misma instancia de DB
        formulario = FormularioInventario(instance=instancia_inicial, ya_existe=ya_existe)

    form_aumentar = FormularioAumentarInventario() if ya_existe else None

    return render(request, "schedule/gestionar_inventario.html", {
        "formulario": formulario,
        "ya_existe": ya_existe,
        "inventario": inventario,
        "form_aumentar": form_aumentar,
    })


@login_required
def aumentar_inventario(request):
    if request.user.role != 'administrador' and not request.user.is_superuser:
        return redirect("dashboard_empleado")

    if request.method == "POST":
        inventario = InventarioTiquetes.objects.order_by("-mes").first()
        if not inventario:
            messages.error(request, "No hay un inventario configurado para aumentar.")
            return redirect("gestionar_inventario")

        form = FormularioAumentarInventario(request.POST)
        if form.is_valid():
            adicion = form.cleaned_data["cantidad_a_adicionar"]
            inventario.cantidad_disponible += adicion
            inventario.cantidad_inicial += adicion # También aumentamos el total del mes para reportes
            inventario.save()
            messages.success(request, f"Se han adicionado {adicion} tiquetes al inventario. Nuevo stock disponible: {inventario.cantidad_disponible}")
        else:
            messages.error(request, "La cantidad a adicionar no es válida.")

    return redirect("gestionar_inventario")


@login_required
def historial_consumos(request, empleado_id):
    if request.user.role != 'administrador' and not request.user.is_superuser:
        return redirect("dashboard_empleado")

    from users.models import Empleado as EmpleadoModel
    empleado_obj = get_object_or_404(EmpleadoModel, id=empleado_id)

    consumos = Consumo.objects.filter(
        empleado_id=empleado_id
    ).order_by('-fecha_consumo')

    return render(request, "schedule/historial_consumo.html", {
        "consumos": consumos,
        "empleado_obj": empleado_obj,
    })


@login_required
def editar_consumo(request, consumo_id):
    if request.user.role != 'administrador' and not request.user.is_superuser:
        return redirect("dashboard_empleado")

    consumo = get_object_or_404(Consumo, id=consumo_id)
    empleado_id_original = consumo.empleado.id

    if request.method == "POST":
        formulario = FormularioEditarConsumo(request.POST, instance=consumo)
        if formulario.is_valid():
            motivo = formulario.cleaned_data.get("motivo", "")

            # Detectar cambios campo a campo y registrarlos en ConsumoLog
            campos_legibles = {
                "empleado": "Empleado",
                "fecha_consumo": "Fecha de consumo",
            }
            
            # Obtener los valores originales de la base de datos
            consumo_original = Consumo.objects.get(id=consumo.id)
            
            for campo in ["empleado", "fecha_consumo"]:
                valor_anterior = getattr(consumo_original, campo)
                valor_nuevo = formulario.cleaned_data.get(campo)
                if valor_anterior != valor_nuevo:
                    ConsumoLog.objects.create(
                        consumo=consumo,
                        editado_por=request.user,
                        campo=campos_legibles.get(campo, campo),
                        valor_anterior=str(valor_anterior),
                        valor_nuevo=str(valor_nuevo),
                        motivo=motivo,
                    )

            formulario.save()
            messages.success(request, "Registro de consumo corregido y cambio registrado en el historial.")
            return redirect("historial_consumos", empleado_id=consumo.empleado.id)
    else:
        formulario = FormularioEditarConsumo(instance=consumo)

    logs = consumo.logs.select_related("editado_por").order_by("-fecha_edicion")

    return render(request, "schedule/editar_consumo.html", {
        "formulario": formulario,
        "consumo": consumo,
        "logs": logs,
        "empleado_id_original": empleado_id_original,
    })


@login_required
def historial_consumos_restaurante(request):
    """Muestra los consumos registrados en una fecha específica (por defecto hoy) por el restaurante."""
    if request.user.role != 'restaurante' and not request.user.is_superuser:
        return redirect("inicio")

    hoy = timezone.localdate()
    fecha_filtro = hoy
    es_hoy = True

    fecha_str = request.GET.get('fecha')
    if fecha_str:
        parsed_date = parse_date(fecha_str)
        if parsed_date:
            fecha_filtro = parsed_date
            es_hoy = (parsed_date == hoy)

    consumos_hoy = Consumo.objects.filter(
        fecha_consumo__date=fecha_filtro
    ).select_related(
        'empleado__user'
    ).order_by('-fecha_consumo')

    return render(request, "schedule/historial_consumos_restaurante.html", {
        "consumos_hoy": consumos_hoy,
        "total_consumos_hoy": consumos_hoy.count(),
        "fecha_filtro": fecha_filtro,
        "es_hoy": es_hoy,
    })


def _usuario_es_admin(user):
    return user.is_superuser or user.role == "administrador"


def _fecha_para_archivo():
    return timezone.localdate().strftime("%Y%m%d")


def _normalizar_formato(formato):
    formato_normalizado = (formato or "csv").strip().lower()
    if formato_normalizado not in {"csv", "excel", "pdf"}:
        return "csv"
    return formato_normalizado


def _normalizar_booleano(valor):
    valor_normalizado = (valor or "").strip().lower()
    if valor_normalizado in {"si", "true", "1"}:
        return True
    if valor_normalizado in {"no", "false", "0"}:
        return False
    return None


def _aplicar_filtros_comunes(queryset, request, campo_fecha):
    empleado_id = (request.GET.get("empleado") or "").strip()
    buscar = (request.GET.get("buscar") or "").strip()
    departamento = (request.GET.get("departamento") or "").strip()
    fecha_desde = parse_date((request.GET.get("fecha_desde") or "").strip())
    fecha_hasta = parse_date((request.GET.get("fecha_hasta") or "").strip())

    if empleado_id.isdigit():
        queryset = queryset.filter(empleado_id=int(empleado_id))

    if buscar:
        queryset = queryset.filter(
            Q(empleado__user__username__icontains=buscar)
            | Q(empleado__user__first_name__icontains=buscar)
            | Q(empleado__user__last_name__icontains=buscar)
            | Q(empleado__numero_documento__icontains=buscar)
            | Q(empleado__departamento__icontains=buscar)
        )

    if departamento:
        queryset = queryset.filter(empleado__departamento=departamento)

    if fecha_desde:
        queryset = queryset.filter(**{f"{campo_fecha}__date__gte": fecha_desde})

    if fecha_hasta:
        queryset = queryset.filter(**{f"{campo_fecha}__date__lte": fecha_hasta})

    return queryset


def _consumos_para_reporte(request=None):
    encabezados = [
        "Empleado",
        "Usuario",
        "Documento",
        "Departamento",
        "Fecha consumo",
    ]
    filas = []

    consumos = Consumo.objects.select_related("empleado__user").order_by("-fecha_consumo")
    if request is not None:
        consumos = _aplicar_filtros_comunes(consumos, request, "fecha_consumo")

    for consumo in consumos:
        empleado = consumo.empleado
        user = empleado.user
        filas.append([
            user.get_full_name() or user.username,
            user.username,
            empleado.numero_documento or "",
            empleado.departamento or "",
            timezone.localtime(consumo.fecha_consumo).strftime("%Y-%m-%d %H:%M:%S"),
        ])

    return encabezados, filas


def _pagos_para_reporte(request=None):
    encabezados = [
        "Empleado",
        "Usuario",
        "Documento",
        "Valor pagado",
        "Fecha pago",
        "Comprobante",
        "Validado por GH",
        "Confirmado por empleado",
    ]
    filas = []

    pagos = RegistroPago.objects.select_related("empleado__user").order_by("-fecha_pago")
    if request is not None:
        pagos = _aplicar_filtros_comunes(pagos, request, "fecha_pago")
        validado = _normalizar_booleano(request.GET.get("validado"))
        confirmado = _normalizar_booleano(request.GET.get("confirmado"))
        if validado is not None:
            pagos = pagos.filter(validado_por_gh=validado)
        if confirmado is not None:
            pagos = pagos.filter(confirmado_por_empleado=confirmado)

    for pago in pagos:
        empleado = pago.empleado
        user = empleado.user
        filas.append([
            user.get_full_name() or user.username,
            user.username,
            empleado.numero_documento or "",
            f"{pago.valor_pagado:.2f}",
            timezone.localtime(pago.fecha_pago).strftime("%Y-%m-%d %H:%M:%S"),
            pago.comprobante or "",
            "Si" if pago.validado_por_gh else "No",
            "Si" if pago.confirmado_por_empleado else "No",
        ])

    return encabezados, filas


def _respuesta_csv(nombre_base, encabezados, filas):
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{nombre_base}.csv"'
    response.write("\ufeff")

    writer = csv.writer(response)
    writer.writerow(encabezados)
    writer.writerows(filas)
    return response


def _respuesta_excel(nombre_base, titulo, encabezados, filas):
    if Workbook is None:
        response = HttpResponse(content_type="application/vnd.ms-excel; charset=utf-8")
        response["Content-Disposition"] = f'attachment; filename="{nombre_base}.xls"'
        response.write("\ufeff")
        response.write("<table border='1'>")
        response.write(f"<tr><th colspan='{len(encabezados)}'>{titulo}</th></tr>")
        response.write("<tr>" + "".join(f"<th>{valor}</th>" for valor in encabezados) + "</tr>")
        for fila in filas:
            response.write("<tr>" + "".join(f"<td>{valor}</td>" for valor in fila) + "</tr>")
        response.write("</table>")
        return response

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Reporte"
    sheet.append([titulo])
    sheet.append(encabezados)

    for fila in filas:
        sheet.append(fila)

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(encabezados))
    sheet["A1"].font = Font(size=14, bold=True, color="16325B")
    sheet["A1"].alignment = Alignment(horizontal="center")

    header_fill = PatternFill(fill_type="solid", start_color="EEF3FF", end_color="EEF3FF")
    for cell in sheet[2]:
        cell.font = Font(bold=True, color="16325B")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for indice, column in enumerate(sheet.columns, start=1):
        max_length = 0
        column_letter = get_column_letter(indice)
        for cell in column:
            cell_value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(cell_value))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 14), 34)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{nombre_base}.xlsx"'
    return response


def _respuesta_pdf(nombre_base, titulo, encabezados, filas):
    if SimpleDocTemplate is None:
        return _respuesta_pdf_basico(nombre_base, titulo, encabezados, filas)

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=36,
        rightMargin=36,
        topMargin=42,
        bottomMargin=36,
    )
    estilos = getSampleStyleSheet()
    elementos = [
        Paragraph(titulo, estilos["Title"]),
        Paragraph(
            f"Generado el {timezone.localtime().strftime('%Y-%m-%d %H:%M:%S')}",
            estilos["BodyText"],
        ),
        Spacer(1, 12),
    ]

    tabla = Table([encabezados] + filas, repeatRows=1)
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EEF3FF")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#16325B")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D6E0F5")),
        ("BACKGROUND", (0, 1), (-1, -1), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elementos.append(tabla)
    doc.build(elementos)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{nombre_base}.pdf"'
    return response


def _respuesta_pdf_basico(nombre_base, titulo, encabezados, filas):
    lineas = [
        titulo,
        f"Generado el {timezone.localtime().strftime('%Y-%m-%d %H:%M:%S')}",
        "",
        " | ".join(encabezados),
        "-" * 120,
    ]
    for fila in filas:
        lineas.append(" | ".join(str(valor) for valor in fila))

    paginas = []
    lineas_por_pagina = 42
    for inicio in range(0, len(lineas), lineas_por_pagina):
        paginas.append(lineas[inicio:inicio + lineas_por_pagina])

    objetos = []

    def _texto_pdf(valor):
        return valor.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")

    for pagina in paginas:
        contenido = ["BT /F1 9 Tf 40 760 Td 12 TL"]
        for indice, linea in enumerate(pagina):
            prefijo = "" if indice == 0 else "T* "
            contenido.append(f"{prefijo}({_texto_pdf(linea[:160])}) Tj")
        contenido.append("ET")
        objetos.append("\n".join(contenido).encode("latin-1", errors="replace"))

    total_paginas = len(objetos)
    font_obj = total_paginas * 2 + 3
    objetos_pdf = [b"<< /Type /Catalog /Pages 2 0 R >>"]
    hijos = " ".join(f"{indice} 0 R" for indice in range(3, 3 + total_paginas * 2, 2))
    objetos_pdf.append(f"<< /Type /Pages /Count {total_paginas} /Kids [{hijos}] >>".encode("ascii"))

    for index, contenido in enumerate(objetos, start=0):
        page_obj = 3 + index * 2
        content_obj = page_obj + 1
        objetos_pdf.append(
            (
                f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] "
                f"/Resources << /Font << /F1 {font_obj} 0 R >> >> /Contents {content_obj} 0 R >>"
            ).encode("ascii")
        )
        objetos_pdf.append(
            f"<< /Length {len(contenido)} >>\nstream\n".encode("ascii") + contenido + b"\nendstream"
        )

    objetos_pdf.append(b"<< /Type /Font /Subtype /Type1 /BaseFont /Courier >>")

    buffer = BytesIO()
    buffer.write(b"%PDF-1.4\n")
    offsets = [0]
    for indice, objeto in enumerate(objetos_pdf, start=1):
        offsets.append(buffer.tell())
        buffer.write(f"{indice} 0 obj\n".encode("ascii"))
        buffer.write(objeto)
        buffer.write(b"\nendobj\n")

    xref_offset = buffer.tell()
    buffer.write(f"xref\n0 {len(objetos_pdf) + 1}\n".encode("ascii"))
    buffer.write(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        buffer.write(f"{offset:010d} 00000 n \n".encode("ascii"))
    buffer.write(
        (
            f"trailer\n<< /Size {len(objetos_pdf) + 1} /Root 1 0 R >>\n"
            f"startxref\n{xref_offset}\n%%EOF"
        ).encode("ascii")
    )

    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{nombre_base}.pdf"'
    return response


def _respuesta_reporte(nombre_base, titulo, encabezados, filas, formato):
    formato_normalizado = _normalizar_formato(formato)
    if formato_normalizado == "excel":
        return _respuesta_excel(nombre_base, titulo, encabezados, filas)
    if formato_normalizado == "pdf":
        return _respuesta_pdf(nombre_base, titulo, encabezados, filas)
    return _respuesta_csv(nombre_base, encabezados, filas)


@login_required
def exportar_reporte_consumos(request):
    if not _usuario_es_admin(request.user):
        return redirect("dashboard_empleado")

    formato = request.GET.get("formato", "csv")
    encabezados, filas = _consumos_para_reporte(request)
    return _respuesta_reporte(
        nombre_base=f"reporte_consumos_{_fecha_para_archivo()}",
        titulo="Reporte de consumos",
        encabezados=encabezados,
        filas=filas,
        formato=formato,
    )


@login_required
def exportar_reporte_pagos(request):
    if not _usuario_es_admin(request.user):
        return redirect("dashboard_empleado")

    formato = request.GET.get("formato", "csv")
    encabezados, filas = _pagos_para_reporte(request)
    return _respuesta_reporte(
        nombre_base=f"reporte_pagos_{_fecha_para_archivo()}",
        titulo="Reporte de pagos",
        encabezados=encabezados,
        filas=filas,
        formato=formato,
    )

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="reporte_pagos_{_fecha_para_archivo()}.csv"'
    response.write("\ufeff")

    writer = csv.writer(response)
    writer.writerow([
        "Empleado",
        "Usuario",
        "Documento",
        "Valor pagado",
        "Fecha pago",
        "Comprobante",
        "Validado por GH",
        "Confirmado por empleado",
    ])

    pagos = RegistroPago.objects.select_related("empleado__user").order_by("-fecha_pago")
    for pago in pagos:
        empleado = pago.empleado
        user = empleado.user
        writer.writerow([
            user.get_full_name() or user.username,
            user.username,
            empleado.numero_documento or "",
            f"{pago.valor_pagado:.2f}",
            timezone.localtime(pago.fecha_pago).strftime("%Y-%m-%d %H:%M:%S"),
            pago.comprobante or "",
            "Sí" if pago.validado_por_gh else "No",
            "Sí" if pago.confirmado_por_empleado else "No",
        ])

    return response


@login_required
def dashboard_reportes(request):
    if request.user.role != 'administrador' and not request.user.is_superuser:
        return redirect("dashboard_empleado")

    empleados = Empleado.objects.select_related("user").order_by("user__first_name", "user__last_name", "user__username")
    departamentos = (
        Empleado.objects.exclude(departamento__isnull=True)
        .exclude(departamento__exact="")
        .values_list("departamento", flat=True)
        .distinct()
        .order_by("departamento")
    )
    filtros = {
        "buscar": request.GET.get("buscar", "").strip(),
        "empleado": request.GET.get("empleado", "").strip(),
        "departamento": request.GET.get("departamento", "").strip(),
        "fecha_desde": request.GET.get("fecha_desde", "").strip(),
        "fecha_hasta": request.GET.get("fecha_hasta", "").strip(),
        "validado": request.GET.get("validado", "").strip(),
        "confirmado": request.GET.get("confirmado", "").strip(),
    }

    return render(
        request,
        "schedule/reportes_admin.html",
        {
            "empleados_filtro": empleados,
            "departamentos_filtro": departamentos,
            "filtros": filtros,
        },
    )
